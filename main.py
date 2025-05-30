import os
import shutil
import re
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from typing import Dict, List, Tuple, Optional, Set
import logging

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def sanitize_filename(name: str) -> str:
    """Clean a string to make it safe for use as a Windows filename."""
    # Replace invalid characters with underscores
    return re.sub(r'[<>:"/\\|?*\n\r]+', '_', name).strip()

def find_table_boundaries(sheet, sheet_name: str) -> Optional[Tuple[int, int, int, int]]:
    """
    Find the data table boundaries in a QA-ID sheet.
    
    Returns:
        Tuple of (header_row, data_start_row, data_end_row, max_col) or None if not found
    """
    try:
        # Step 1: Find "Detailed Results" in column B
        detailed_results_row = None
        for row_num in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_num, column=2).value
            if cell_value and "Detailed Results" in str(cell_value):
                detailed_results_row = row_num
                logger.info(f"Found 'Detailed Results' at row {row_num} in {sheet_name}")
                break
        
        if detailed_results_row is None:
            logger.warning(f"Could not find 'Detailed Results' in column B for sheet {sheet_name}")
            return None
        
        # Step 2: Find "Audit Leader" in column B after "Detailed Results"
        audit_leader_row = None
        for row_num in range(detailed_results_row + 1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_num, column=2).value
            if cell_value and "Audit Leader" in str(cell_value):
                audit_leader_row = row_num
                logger.info(f"Found 'Audit Leader' at row {row_num} in {sheet_name}")
                break
        
        if audit_leader_row is None:
            logger.warning(f"Could not find 'Audit Leader' in column B after 'Detailed Results' for sheet {sheet_name}")
            return None
        
        # Step 3: Determine data boundaries
        header_row = audit_leader_row
        data_start_row = header_row + 1
        
        # Find the end of data by looking for first completely empty row
        data_end_row = sheet.max_row
        for row_num in range(data_start_row, sheet.max_row + 1):
            row_has_data = False
            for col_num in range(1, sheet.max_column + 1):
                if sheet.cell(row=row_num, column=col_num).value is not None:
                    row_has_data = True
                    break
            if not row_has_data:
                data_end_row = row_num - 1
                break
        
        # Find the actual max column used in the header row
        max_col = 1
        for col_num in range(1, sheet.max_column + 1):
            if sheet.cell(row=header_row, column=col_num).value is not None:
                max_col = col_num
        
        logger.info(f"Table boundaries for {sheet_name}: header_row={header_row}, data_start={data_start_row}, data_end={data_end_row}, max_col={max_col}")
        return header_row, data_start_row, data_end_row, max_col
        
    except Exception as e:
        logger.error(f"Error finding table boundaries in {sheet_name}: {str(e)}")
        return None

def get_column_mapping(sheet, header_row: int, max_col: int) -> Dict[str, int]:
    """
    Create a mapping of column names to column numbers from the header row.
    """
    column_mapping = {}
    for col_num in range(1, max_col + 1):
        header_value = sheet.cell(row=header_row, column=col_num).value
        if header_value:
            column_mapping[str(header_value).strip()] = col_num
    
    logger.info(f"Column mapping: {column_mapping}")
    return column_mapping

def get_cell_calculated_value(cell):
    """
    Get the calculated value of a cell, handling formulas properly.
    """
    if cell.data_type == 'f':  # Formula cell
        # Return the cached calculated value
        return cell.value
    else:
        # Regular value
        return cell.value

def extract_data_to_dataframe_with_calculated_values(wb_calculated, sheet_name: str, header_row: int, 
                                                   data_start_row: int, data_end_row: int, max_col: int) -> pd.DataFrame:
    """
    Extract data from the sheet into a pandas DataFrame using calculated values.
    """
    sheet_calc = wb_calculated[sheet_name]
    
    # Get headers
    headers = []
    for col_num in range(1, max_col + 1):
        header_value = sheet_calc.cell(row=header_row, column=col_num).value
        headers.append(str(header_value) if header_value is not None else f"Column_{col_num}")
    
    # Get data with calculated values
    data = []
    for row_num in range(data_start_row, data_end_row + 1):
        row_data = []
        for col_num in range(1, max_col + 1):
            cell_value = get_cell_calculated_value(sheet_calc.cell(row=row_num, column=col_num))
            row_data.append(cell_value)
        data.append(row_data)
    
    df = pd.DataFrame(data, columns=headers)
    logger.info(f"Extracted DataFrame with calculated values, shape {df.shape}")
    return df

def finalize_sheet_presentation(sheet):
    """
    Set sheet to A1 position and collapse all grouped rows/columns before saving.
    """
    try:
        # Set sheet view to show A1 (top-left corner)
        sheet.sheet_view.topLeftCell = "A1"
        
        # Set the selection to A1 (proper way to set active cell)
        if hasattr(sheet.sheet_view, 'selection') and len(sheet.sheet_view.selection) > 0:
            sheet.sheet_view.selection[0].activeCell = "A1"
            sheet.sheet_view.selection[0].sqref = "A1"
        
        # Collapse all row groups (outline levels)
        if hasattr(sheet, 'row_dimensions'):
            for row_num, row_dim in sheet.row_dimensions.items():
                if hasattr(row_dim, 'outline_level') and row_dim.outline_level > 0:
                    row_dim.hidden = True
        
        # Collapse all column groups (outline levels)  
        if hasattr(sheet, 'column_dimensions'):
            for col_letter, col_dim in sheet.column_dimensions.items():
                if hasattr(col_dim, 'outline_level') and col_dim.outline_level > 0:
                    col_dim.hidden = True
        
        # Set outline summary below/right to collapse groups
        if hasattr(sheet, 'sheet_properties') and hasattr(sheet.sheet_properties, 'outline_pr'):
            sheet.sheet_properties.outline_pr.summary_below = False
            sheet.sheet_properties.outline_pr.summary_right = False
        
        # Clear any existing freeze panes
        sheet.freeze_panes = None
        
        logger.info(f"Finalized presentation for sheet: {sheet.title}")
        
    except Exception as e:
        logger.warning(f"Could not finalize presentation for sheet {sheet.title}: {str(e)}")

def get_result_column_number(sheet, header_row: int, max_col: int) -> Optional[int]:
    """Find the result column number using the same logic as before."""
    target_text = "overall test result (after considering any applicable test result overrides)"
    
    for col_num in range(1, max_col + 1):
        header_value = sheet.cell(row=header_row, column=col_num).value
        if header_value:
            normalized_col = ' '.join(str(header_value).replace('\n', ' ').split()).lower()
            if normalized_col == target_text:
                return col_num
    
    # Fallback logic
    for col_num in range(1, max_col + 1):
        header_value = sheet.cell(row=header_row, column=col_num).value
        if header_value:
            normalized_col = ' '.join(str(header_value).replace('\n', ' ').split()).lower()
            if ("overall test result" in normalized_col and 
                "considering" in normalized_col and 
                "applicable" in normalized_col and
                "override" not in normalized_col):
                return col_num
    
    return None

def get_audit_leader_column_number(column_mapping: Dict[str, int]) -> int:
    """Find the column number for Audit Leader."""
    for col_name, col_num in column_mapping.items():
        if "Audit Leader" in col_name:
            return col_num
    raise ValueError("Audit Leader column not found in column mapping")

def sort_sheet_by_audit_leader_and_dnc(sheet_formula, sheet_calculated, header_row: int, data_start_row: int, 
                                      data_end_row: int, audit_leader_col: int, result_col_num: int = None):
    """
    Sort all data in the sheet by audit leader first, then DNC within each leader.
    Uses calculated values for sorting but writes back to formula sheet.
    """
    logger.info(f"Sorting sheet by audit leader and DNC using calculated values...")
    
    # Read all row data with sort keys from both sheets
    all_rows_with_keys = []
    
    for row_num in range(data_start_row, data_end_row + 1):
        # Read entire row data from formula sheet (what we'll write back)
        row_data = []
        for col_num in range(1, sheet_formula.max_column + 1):
            cell_value = sheet_formula.cell(row=row_num, column=col_num).value
            row_data.append(cell_value)
        
        # Get audit leader from calculated sheet for sorting
        audit_leader = get_cell_calculated_value(sheet_calculated.cell(row=row_num, column=audit_leader_col))
        audit_leader_str = str(audit_leader or "").strip()
        
        # Get DNC status from calculated sheet for secondary sorting
        has_dnc = False
        if result_col_num:
            result_value = get_cell_calculated_value(sheet_calculated.cell(row=row_num, column=result_col_num))
            has_dnc = result_value and "DNC" in str(result_value).upper()
        
        # Create sort key: (audit_leader, not has_dnc)
        sort_key = (audit_leader_str, not has_dnc)
        
        all_rows_with_keys.append((sort_key, row_data))
    
    # Sort by the sort key
    all_rows_with_keys.sort(key=lambda x: x[0])
    
    # Extract just the row data in sorted order
    sorted_data = [row_data for sort_key, row_data in all_rows_with_keys]
    
    # Write sorted data back to formula sheet
    for idx, row_data in enumerate(sorted_data):
        excel_row = data_start_row + idx
        for col_idx, value in enumerate(row_data):
            sheet_formula.cell(row=excel_row, column=col_idx + 1).value = value
    
    # Create summary of the sort
    audit_leader_counts = {}
    for sort_key, _ in all_rows_with_keys:
        leader = sort_key[0]
        audit_leader_counts[leader] = audit_leader_counts.get(leader, 0) + 1
    
    logger.info(f"Sorted {len(sorted_data)} rows by audit leader:")
    for leader, count in sorted(audit_leader_counts.items()):
        if count > 0:
            logger.info(f"  {leader}: {count} rows")

def find_audit_leader_boundaries(sheet_calculated, data_start_row: int, data_end_row: int, 
                                audit_leader_col: int, target_leader: str) -> Tuple[Optional[int], Optional[int]]:
    """
    Find the start and end rows for a specific audit leader using calculated values.
    """
    start_row = None
    end_row = None
    
    for row_num in range(data_start_row, data_end_row + 1):
        audit_leader = get_cell_calculated_value(sheet_calculated.cell(row=row_num, column=audit_leader_col))
        audit_leader_str = str(audit_leader or "").strip()
        
        if audit_leader_str == target_leader:
            if start_row is None:
                start_row = row_num
            end_row = row_num
        elif start_row is not None:
            break
    
    if start_row is None:
        logger.warning(f"No rows found for audit leader: {target_leader}")
        return None, None
    
    logger.info(f"Found {target_leader} in rows {start_row} to {end_row} ({end_row - start_row + 1} rows)")
    return start_row, end_row

def filter_sheet_by_bulk_delete(sheet_formula, sheet_calculated, audit_leader: str, data_start_row: int, 
                               data_end_row: int, audit_leader_col: int, result_col_num: int = None) -> bool:
    """
    Filter sheet using bulk deletions after sorting, using calculated values for decisions.
    """
    # Find where this audit leader's data is located using calculated values
    leader_start, leader_end = find_audit_leader_boundaries(
        sheet_calculated, data_start_row, data_end_row, audit_leader_col, audit_leader
    )
    
    if leader_start is None:
        # No rows for this audit leader - delete all data
        total_rows = data_end_row - data_start_row + 1
        if total_rows > 0:
            sheet_formula.delete_rows(data_start_row, total_rows)
            logger.info(f"No data found for {audit_leader} - deleted all {total_rows} data rows")
        return False
    
    # Check for DNC values in the remaining data using calculated values
    has_dnc = False
    if result_col_num:
        for row_num in range(leader_start, leader_end + 1):
            result_value = get_cell_calculated_value(sheet_calculated.cell(row=row_num, column=result_col_num))
            if result_value and "DNC" in str(result_value).upper():
                has_dnc = True
                logger.info(f"Found DNC value at row {row_num}: {result_value}")
                break
    
    # Perform bulk deletions on formula sheet
    deletions_made = 0
    
    # Delete everything after this audit leader's data
    rows_after = data_end_row - leader_end
    if rows_after > 0:
        sheet_formula.delete_rows(leader_end + 1, rows_after)
        deletions_made += rows_after
        logger.info(f"Deleted {rows_after} rows after {audit_leader}'s data")
    
    # Delete everything before this audit leader's data
    rows_before = leader_start - data_start_row
    if rows_before > 0:
        sheet_formula.delete_rows(data_start_row, rows_before)
        deletions_made += rows_before
        logger.info(f"Deleted {rows_before} rows before {audit_leader}'s data")
    
    remaining_rows = (leader_end - leader_start + 1)
    logger.info(f"Filtering complete for {audit_leader}: {remaining_rows} rows remain, "
               f"{deletions_made} rows deleted, DNC present: {has_dnc}")
    
    return has_dnc

def analyze_workbook_structure(workbook_path: str) -> Tuple[Set[str], Dict[str, Tuple]]:
    """
    Analyze workbook structure using calculated values to get audit leaders and table boundaries.
    """
    audit_leaders = set()
    sheet_info = {}
    
    try:
        # Load with calculated values for analysis
        wb_calculated = openpyxl.load_workbook(workbook_path, data_only=True)
        
        for sheet_name in wb_calculated.sheetnames:
            if sheet_name.startswith("QA-ID-"):
                logger.info(f"Analyzing sheet {sheet_name}")
                sheet = wb_calculated[sheet_name]
                
                # Find table boundaries
                boundaries = find_table_boundaries(sheet, sheet_name)
                if boundaries is None:
                    continue
                
                header_row, data_start_row, data_end_row, max_col = boundaries
                column_mapping = get_column_mapping(sheet, header_row, max_col)
                
                # Extract data using calculated values
                df = extract_data_to_dataframe_with_calculated_values(
                    wb_calculated, sheet_name, header_row, data_start_row, data_end_row, max_col
                )
                
                # Find the correct result column
                result_col_name = None
                target_text = "overall test result (after considering any applicable test result overrides)"
                
                for col_name in df.columns:
                    normalized_col = ' '.join(str(col_name).replace('\n', ' ').split()).lower()
                    if normalized_col == target_text:
                        result_col_name = col_name
                        logger.info(f"Found target result column in {sheet_name}: '{col_name}'")
                        break
                
                # Fallback search
                if result_col_name is None:
                    for col_name in df.columns:
                        normalized_col = ' '.join(str(col_name).replace('\n', ' ').split()).lower()
                        if ("overall test result" in normalized_col and 
                            "considering" in normalized_col and 
                            "applicable" in normalized_col and
                            "override" not in normalized_col):
                            result_col_name = col_name
                            logger.info(f"Found fallback result column in {sheet_name}: '{col_name}'")
                            break
                
                if result_col_name is None:
                    logger.warning(f"Could not find result column in {sheet_name}")
                
                # Store sheet info
                sheet_info[sheet_name] = (header_row, data_start_row, data_end_row, max_col, column_mapping, result_col_name)
                
                # Find Audit Leader column and extract unique values using calculated values
                audit_leader_col = None
                for col_name, col_num in column_mapping.items():
                    if "Audit Leader" in col_name:
                        audit_leader_col = col_name
                        break
                
                if audit_leader_col is None:
                    logger.warning(f"Could not find Audit Leader column in {sheet_name}")
                    continue
                
                # Extract unique audit leaders using calculated values
                for leader_value in df[audit_leader_col].dropna().unique():
                    if str(leader_value).strip():
                        audit_leaders.add(str(leader_value).strip())
        
        wb_calculated.close()
        
    except Exception as e:
        logger.error(f"Error analyzing workbook structure: {str(e)}")
    
    logger.info(f"Found unique audit leaders: {sorted(audit_leaders)}")
    logger.info(f"Analyzed {len(sheet_info)} QA-ID sheets")
    return audit_leaders, sheet_info

def create_presorted_workbook(source_file: str, audit_leaders: set, sheet_info: dict) -> str:
    """
    Create a workbook sorted by audit leader, then DNC using calculated values for sorting.
    """
    source_path = Path(source_file)
    sorted_filename = f"{source_path.stem}_sorted_by_leader{source_path.suffix}"
    sorted_path = source_path.parent / sorted_filename
    
    # Copy original file
    shutil.copyfile(source_file, sorted_path)
    logger.info(f"Created pre-sort copy: {sorted_path}")
    
    # Open both formula and calculated versions
    wb_formula = openpyxl.load_workbook(sorted_path, data_only=False)
    wb_calculated = openpyxl.load_workbook(sorted_path, data_only=True)
    
    for sheet_name, (header_row, data_start_row, data_end_row, max_col, column_mapping, result_col_name) in sheet_info.items():
        logger.info(f"Sorting sheet by audit leader: {sheet_name}")
        sheet_formula = wb_formula[sheet_name]
        sheet_calculated = wb_calculated[sheet_name]
        
        # Get column numbers
        audit_leader_col_num = get_audit_leader_column_number(column_mapping)
        result_col_num = get_result_column_number(sheet_formula, header_row, max_col)
        
        # Sort by audit leader, then DNC using calculated values
        sort_sheet_by_audit_leader_and_dnc(
            sheet_formula, sheet_calculated, header_row, data_start_row, data_end_row,
            audit_leader_col_num, result_col_num
        )
    
    wb_formula.save(sorted_path)
    wb_formula.close()
    wb_calculated.close()
    logger.info(f"Pre-sorted workbook saved: {sorted_path}")
    
    return str(sorted_path)

def process_workbook_by_audit_leaders(source_file: str, output_dir: str = None) -> Dict[str, str]:
    """
    Process an Excel workbook to create filtered versions for each audit leader.
    Now properly handles formula cells by using calculated values for decisions.
    """
    source_path = Path(source_file)
    if not source_path.exists():
        raise FileNotFoundError(f"Source file not found: {source_file}")
    
    if output_dir is None:
        output_dir = source_path.parent
    else:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
    
    base_name = source_path.stem
    
    # Step 1: Analyze workbook structure using calculated values
    logger.info("Analyzing workbook structure with calculated values...")
    audit_leaders, sheet_info = analyze_workbook_structure(source_file)
    
    if not audit_leaders:
        logger.warning("No audit leaders found in the workbook")
        return {}
    
    if not sheet_info:
        logger.warning("No QA-ID sheets with valid table structure found")
        return {}
    
    # Step 2: Create pre-sorted workbook
    logger.info("Pre-sorting workbook by audit leader and DNC values...")
    sorted_workbook_path = create_presorted_workbook(source_file, audit_leaders, sheet_info)
    
    results = {}
    
    # Step 3: Process each audit leader
    for audit_leader in sorted(audit_leaders):
        logger.info(f"Processing workbook for audit leader: {audit_leader}")
        
        try:
            # Copy the pre-sorted file
            safe_leader_name = sanitize_filename(audit_leader)
            output_filename = f"{base_name} - {safe_leader_name}.xlsx"
            output_path = output_dir / output_filename
            shutil.copyfile(sorted_workbook_path, output_path)
            logger.info(f"Created copy from pre-sorted file: {output_path}")
            
            # Open both formula and calculated versions
            wb_formula = openpyxl.load_workbook(output_path, data_only=False)
            wb_calculated = openpyxl.load_workbook(output_path, data_only=True)
            
            # Disable auto-calculation for speed
            try:
                wb_formula.calculation.calcMode = 'manual'
                logger.info("Disabled automatic calculation for faster processing")
            except Exception as e:
                logger.warning(f"Could not disable calculation mode: {e}")
            
            # Process each QA-ID sheet
            for sheet_name, (header_row, data_start_row, data_end_row, max_col, column_mapping, result_col_name) in sheet_info.items():
                logger.info(f"Filtering sheet: {sheet_name}")
                sheet_formula = wb_formula[sheet_name]
                sheet_calculated = wb_calculated[sheet_name]
                
                # Get column numbers
                audit_leader_col_num = get_audit_leader_column_number(column_mapping)
                result_col_num = get_result_column_number(sheet_formula, header_row, max_col)
                
                # Filter using calculated values for decisions
                has_dnc = filter_sheet_by_bulk_delete(
                    sheet_formula, sheet_calculated, audit_leader, data_start_row, data_end_row, 
                    audit_leader_col_num, result_col_num
                )
                
                # Set tab color based on DNC presence
                if has_dnc:
                    sheet_formula.sheet_properties.tabColor = "FF0000"  # Red
                    logger.info(f"Set {sheet_name} tab color to red (DNC values present)")
                else:
                    sheet_formula.sheet_properties.tabColor = "00FF00"  # Green
                    logger.info(f"Set {sheet_name} tab color to green (no DNC values)")
            
            # Re-enable calculation
            try:
                wb_formula.calculation.calcMode = 'automatic'
                logger.info("Re-enabled automatic calculation")
            except Exception as e:
                logger.warning(f"Could not re-enable calculation mode: {e}")
            
            # Finalize all sheets
            for sheet_name in wb_formula.sheetnames:
                sheet = wb_formula[sheet_name]
                finalize_sheet_presentation(sheet)
            
            # Save and close
            wb_formula.save(output_path)
            wb_formula.close()
            wb_calculated.close()
            
            results[audit_leader] = str(output_path)
            logger.info(f"Successfully processed workbook for {audit_leader}")
            
        except Exception as e:
            logger.error(f"Error processing workbook for {audit_leader}: {str(e)}")
            if output_path.exists():
                output_path.unlink()
    
    # Clean up temporary file
    try:
        Path(sorted_workbook_path).unlink()
        logger.info("Cleaned up temporary pre-sorted file")
    except Exception as e:
        logger.warning(f"Could not clean up temporary file {sorted_workbook_path}: {str(e)}")
    
    logger.info(f"Processing complete. Created {len(results)} workbooks.")
    return results

# Example usage
if __name__ == "__main__":
    source_file = "your_workbook.xlsx"
    output_directory = "output_files"
    
    try:
        results = process_workbook_by_audit_leaders(source_file, output_directory)
        
        print("Processing Results:")
        for audit_leader, output_file in results.items():
            print(f"  {audit_leader}: {output_file}")
            
    except Exception as e:
        print(f"Error: {e}")
